import os
from werkzeug.utils import secure_filename

from flask import Flask, render_template,request,redirect,flash,session

from reportlab.pdfgen import canvas
from flask import send_file

from openpyxl import Workbook

import sqlite3

app=Flask(__name__)
app.secret_key="hospital123"
    
@app.route("/patients")
def patients():
    conn=sqlite3.connect("database.db")
    cursor=conn.cursor()
    
    cursor.execute("SELECT * FROM patients")
    patients=cursor.fetchall()
    
    print("Patients=",patients)
    
    conn.close()
        
    return render_template("patients.html",patients=patients)

@app.route("/edit_patient/<int:id>",methods=["GET","POST"])
def edit_patient(id):
    conn=sqlite3.connect("database.db")
    cursor=conn.cursor()
    
    cursor.execute("SELECT * FROM patients WHERE id=?",(id,))
    patient=cursor.fetchone()
    
    if request.method == "POST":
        name=request.form["name"]
        age=request.form["age"]
        gender=request.form["gender"]
        phone=request.form["phone"]
        disease=request.form["disease"]
        photo=request.files["photo"]
        filename=patient[6]
        
        if photo and photo.filename !="":
            filename=secure_filename(photo.filename)
            photo.save(os.path.join("static","uploads",filename))
        
        cursor.execute("""
        UPDATE patients SET name=?,age=?,gender=?,phone=?,disease=?,photo=?
        WHERE id=?
        """,(name,age,gender,phone,disease,filename,id))
        
        conn.commit()
        conn.close()
        
        flash("✏️ Patient Updated Successfully")
        
        return redirect("/patients")
    
    conn.close()
    return render_template("edit_patient.html",patient=patient)

@app.route("/delete_patient/<int:id>")
def delete_patient(id):
    conn=sqlite3.connect("database.db")
    cursor=conn.cursor()
    
    cursor.execute("DELETE FROM patients WHERE id=?",(id,))
    conn.commit()
    conn.close()
    
    flash("🗑️ Patient Deleted Successfully")
    
    return redirect("/patients")

@app.route("/add_patient",methods=["GET","POST"])
def add_patient():
    
    if request.method=="POST":
        
        name=request.form["name"]
        age=request.form["age"]
        gender=request.form["gender"]
        phone=request.form["phone"]
        disease=request.form["disease"]
        
        from datetime import datetime
        
        created_at=datetime.now().strftime("%Y-%m-%d")
        
        photo=request.files["photo"]
        filename=""
        
        if photo and photo.filename !="":
            filename=secure_filename(photo.filename)
            photo.save(os.path.join("static","uploads",filename))
            
        conn=sqlite3.connect("database.db")
        cursor=conn.cursor()
        
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS patients(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT,
        age INTEGER,
        gender TEXT,
        phone TEXT,
        disease TEXT,
        photo TEXT,
        created_at TEXT
        )
        """)
        
        try:
            cursor.execute("ALTER TABLE patients ADD COLUMN photo TEXT")
            conn.commit()
        except:
            pass
        
        try:
            cursor.execute("ALTER TABLE patients ADD COLUMN created_at TEXT")
            conn.commit()
        except sqlite3.OperationalError:
            pass
        
        cursor.execute("""
        INSERT INTO patients(name,age,gender,phone,disease,photo,created_at)
        VALUES(?,?,?,?,?,?,?)
        """,(name,age,gender,phone,disease,filename,created_at))
        
        conn.commit()
        conn.close()
        
        flash("✅ Patient Added Successfully")
  
        return redirect("/patients")

    return render_template("add_patient.html")

@app.route("/doctors")
def doctors():
    conn=sqlite3.connect("database.db")
    cursor=conn.cursor()
    
    cursor.execute("SELECT * FROM doctors")
    doctors=cursor.fetchall()
    
    conn.close()
        
    return render_template("doctors.html",doctors=doctors)

@app.route("/add_doctor", methods=["GET", "POST"])
def add_doctor():

    if request.method == "POST":

        name = request.form["name"]
        specialization = request.form["specialization"]
        phone = request.form["phone"]
        experience = request.form["experience"]

        photo = request.files["photo"]
        filename = ""

        if photo and photo.filename != "":
            filename = secure_filename(photo.filename)
            photo.save(os.path.join("static", "uploads", filename))

        conn = sqlite3.connect("database.db")
        cursor = conn.cursor()

        cursor.execute("""
        CREATE TABLE IF NOT EXISTS doctors(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT,
            specialization TEXT,
            phone TEXT,
            experience INTEGER,
            photo TEXT
        )
        """)

        try:
            cursor.execute("ALTER TABLE doctors ADD COLUMN photo TEXT")
        except sqlite3.OperationalError:
            pass

        cursor.execute("""
        INSERT INTO doctors(name, specialization, phone, experience, photo)
        VALUES (?, ?, ?, ?, ?)
        """, (name, specialization, phone, experience, filename))

        conn.commit()
        conn.close()

        flash("✅ Doctor Added Successfully")

        return redirect("/doctors")

    return render_template("add_doctor.html")

@app.route("/edit_doctor/<int:id>",methods=["GET","POST"])
def edit_doctor(id):
     conn=sqlite3.connect("database.db")
     cursor=conn.cursor()
     
     try:
         cursor.execute("ALTER TABLE doctors ADD COLUMN photo TEXT")
         conn.commit()
     except sqlite3.OperationalError:
         pass
     
     cursor.execute("PRAGMA table_info(doctors)")
     print(cursor.fetchall())
     
     cursor.execute("SELECT * FROM doctors WHERE id=?",(id,))
     doctor = cursor.fetchone()
     print(doctor)
     
     if request.method == "POST":
        name=request.form["name"]
        specialization=request.form["specialization"]
        phone=request.form["phone"]
        experience=request.form["experience"]
        
        photo=request.files["photo"]
        if len(doctor)>5:
            filename=doctor[5]
        else:
            filename=""
        
        if photo and photo.filename !="":
            filename=secure_filename(photo.filename)
            photo.save(os.path.join("static","uploads",filename))
        
        cursor.execute("""
        UPDATE doctors SET name=?,specialization=?,phone=?,experience=?,photo=?
        WHERE id=?
        """,(name,specialization,phone,experience,filename,id))
        
        conn.commit()
        conn.close()
        
        return redirect("/doctors")
    
     conn.close()
    
     return render_template("edit_doctor.html",doctor=doctor)
 
@app.route("/delete_doctor/<int:id>")
def delete_doctor(id):
    conn=sqlite3.connect("database.db")
    cursor=conn.cursor()
    
    cursor.execute("DELETE FROM doctors WHERE id=?",(id,))
    conn.commit()
    conn.close()
    return redirect("/doctors")

@app.route("/appointments")
def appointments():
    conn=sqlite3.connect("database.db")
    cursor=conn.cursor()
    
    cursor.execute("SELECT * FROM appointments")
    appointments=cursor.fetchall()
    
    print("Appointments =",appointments)
    
    conn.close()
    
    return render_template("appointments.html",appointments=appointments)

@app.route("/add_appointment",methods=["GET","POST"])
def add_appointment():
    
    if request.method == "POST":
        patient_name = request.form["patient_name"]
        doctor_name = request.form["doctor_name"]
        date=request.form["date"]
        time=request.form["time"]
        status=request.form["status"]
        
        print(patient_name,doctor_name,date,time)
        
        conn=sqlite3.connect("database.db")
        cursor=conn.cursor()
        
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS appointments(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            patient_name TEXT,
            doctor_name TEXT,
            date TEXT,
            time TEXT,
            status TEXT
        )
        """)
        
        cursor.execute("""
        INSERT INTO appointments(patient_name,doctor_name,date,time,status)
        VALUES(?,?,?,?,?)
        """,(patient_name,doctor_name,date,time,status))
        
        conn.commit()
        conn.close()
        
        return redirect("/appointments")
    
    return render_template("add_appointment.html")

@app.route("/edit_appointment/<int:id>",methods=["GET","POST"])
def edit_appointment(id):
    conn=sqlite3.connect("database.db")
    cursor=conn.cursor()
    
    if request.method=="POST":
        patient_name=request.form["patient_name"]
        doctor_name=request.form["doctor_name"]
        date=request.form["date"]
        time=request.form["time"]
        status=request.form["status"]
        
        cursor.execute("""
        UPDATE appointments
        SET patient_name=?, doctor_name=?, date=?, time=?, status=?
        WHERE id=?
        """, (patient_name, doctor_name, date, time, status, id))
        
        conn.commit()
        conn.close()

        return redirect("/appointments")
    
    cursor.execute("SELECT * FROM appointments WHERE id=?",(id,))
    appointment=cursor.fetchone()
    
    conn.close()
    
    return render_template("edit_appointment.html",appointment=appointment)

@app.route("/delete_appointment/<int:id>")
def delete_appointment(id):
    conn=sqlite3.connect("database.db")
    cursor=conn.cursor()
    
    cursor.execute("DELETE FROM appointments WHERE id=?",(id,))
    
    conn.commit()
    conn.close()
    
    return redirect("/appointments")

@app.route("/login",methods=["GET","POST"])
def login():
    
    if request.method=="POST":
        
        username=request.form["username"]
        password=request.form["password"]
        
        if username=="admin" and password=="admin123":
            session["user"]=username
            return redirect("/")
        
        flash("Invalid Username or Password")
        
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.pop("user",None)
    return redirect("/login")

@app.route("/")
def home():
    
    if "user" not in session:
        return redirect("/login")
    
    conn=sqlite3.connect("database.db")
    cursor=conn.cursor()
    
    cursor.execute("SELECT COUNT(*) FROM patients")
    patient_count=cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(*) FROM doctors")
    doctor_count=cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(*) FROM appointments")
    appointment_count=cursor.fetchone()[0]
    
    from datetime import datetime
    
    today=datetime.now().strftime("%Y-%m-%d")
    
    cursor.execute("SELECT COUNT(*) FROM patients WHERE created_at=?",(today,))
    today_patients=cursor.fetchone()[0]
    
    cursor.execute("""
    SELECT patient_name, doctor_name, date, time, status
    FROM appointments
    ORDER BY id DESC
    LIMIT 5
    """)

    recent_appointments = cursor.fetchall()
    
    conn.close()
    
    return render_template(
    "index.html",
    patient_count=patient_count,
    doctor_count=doctor_count,
    appointment_count=appointment_count,
    today_patients=today_patients,
    recent_appointments=recent_appointments
    )
    
@app.route("/reports")
def reports():
    if "user" not in session:
        return redirect("/login")

    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    cursor.execute("SELECT COUNT(*) FROM patients")
    patient_count = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM doctors")
    doctor_count = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM appointments")
    appointment_count = cursor.fetchone()[0]

    conn.close()

    return render_template(
        "reports.html",
        patient_count=patient_count,
        doctor_count=doctor_count,
        appointment_count=appointment_count
    )
    
@app.route("/settings")
def settings():
    if "user" not in session:
        return redirect("/login")

    return render_template("settings.html")

@app.route("/download_report")
def download_report():

    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    cursor.execute("SELECT COUNT(*) FROM patients")
    patients = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM doctors")
    doctors = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM appointments")
    appointments = cursor.fetchone()[0]

    conn.close()

    pdf = canvas.Canvas("Hospital_Report.pdf")

    pdf.setFont("Helvetica-Bold",18)
    pdf.drawString(180,800,"Hospital Report")

    pdf.setFont("Helvetica",14)
    pdf.drawString(100,740,f"Total Patients : {patients}")
    pdf.drawString(100,710,f"Total Doctors : {doctors}")
    pdf.drawString(100,680,f"Total Appointments : {appointments}")

    pdf.save()

    return send_file("Hospital_Report.pdf",as_attachment=True)

@app.route("/download_excel")
def download_excel():

    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    wb = Workbook()

    # ---------------- Patients ----------------
    ws1 = wb.active
    ws1.title = "Patients"

    ws1.append(["ID", "Name", "Age", "Gender", "Phone", "Disease"])

    cursor.execute("SELECT * FROM patients")
    for row in cursor.fetchall():
        ws1.append(row)

    # ---------------- Doctors ----------------
    ws2 = wb.create_sheet(title="Doctors")

    ws2.append(["ID", "Name", "Specialization", "Phone", "Experience"])

    cursor.execute("SELECT * FROM doctors")
    for row in cursor.fetchall():
        ws2.append(row)

    # ---------------- Appointments ----------------
    ws3 = wb.create_sheet(title="Appointments")

    ws3.append(["ID", "Patient", "Doctor", "Date", "Time"])

    cursor.execute("SELECT * FROM appointments")
    for row in cursor.fetchall():
        ws3.append(row)

    conn.close()

    filename = "Hospital_Report.xlsx"
    wb.save(filename)

    return send_file(filename, as_attachment=True)


if __name__=="__main__":
    app.secret_key = "hospital123"
    
    app.run(debug=True)

